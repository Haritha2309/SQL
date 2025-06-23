# 1. text file handling

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\sample.txt","w") as file:
    file.write("File handling started with text file. First create a new folder and a new blank text file. Then add comments to it in this format")

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\sample.txt","a") as file:
    file.write("\n\nThen I understood that new blank file is not necessary. Just copy path and new file will be opened when u give name in file.open commend")

with open ("C:\\Users\\Haritha\\Desktop\\python\\practise\\sample.txt","r") as file:
    print(file.read())

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\sample.txt","a") as file:
    file.write("\n\nWith this we came to an end that only we write in vs code will be displayed on terminal and not everything we write on text document. But if needed to add lines we should write those lines in vs code only")

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\sample.txt","a") as file:
    file.write("\n\nEven when we save the text written ton text doc it will not be saved and only what we don in vs code is saved and displayed at further operations")


# 2. file handling xcel

from openpyxl import Workbook,load_workbook
wb= Workbook()
sheet=wb.active
sheet.title = 'xcel_file_handling'
sheet["A1"] ="NAME"
sheet["B1"]="AGE"
sheet.append(["HARITHA",21])
sheet.append(["HARI",29])
wb.save("C:\\Users\\Haritha\\Desktop\\python\\practise\\excelpractise.xlsx")
wb=load_workbook("C:\\Users\\Haritha\\Desktop\\python\\practise\\excelpractise.xlsx")
for row in sheet.iter_rows(min_row=2,values_only = True):
    name,age = row
    print(f"Name:{name},:Age:{age}")

# 3. FILE HANDLING CSV

import csv
data = [['Name','Age'],['Haritha',22],['Hari',25],['Lokesh',24]]
with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\csvpractise.csv","w") as file:
    writer = csv.writer(file)
    writer.writerows(data)

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\csvpractise.csv","r") as file:
    read = csv.reader(file)
    for value in read:
        print(value)


# 4. File handling xml

from bs4 import BeautifulSoup
xml_data = """
    <Student>
        <student_id = "1">
            <student_name> "HARITHA" </student_name>
            <stu_hobby> "Sleep" </stu_hobby>
            <stu_code> "Unique" </stu_code>
"""
soup = BeautifulSoup(xml_data,'xml')
with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\xmlpractise.xml","w") as file:
    file.write(soup.prettify())

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\xmlpractise.xml","r") as file:
    soup = BeautifulSoup(xml_data,'xml')
    hobby = soup.find("stu_code").text
    print(hobby)


# 5. File handling json

import json
data = {
    "person1":{
        "Name": "Haritha",
        "Age":22
    },
    "person2":{
        "Name":"Hari",
        "Age": 25
    }
}
with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\jsonpractise.json","w") as file:
    json.dump(data,file,indent=8)

with open("C:\\Users\\Haritha\\Desktop\\python\\practise\\jsonpractise.json","r") as file:
    json.load(file)
    for i,j in data["person1"].items():
        print(f"{i}-->{j}")